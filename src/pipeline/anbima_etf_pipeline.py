"""
anbima_etf_pipeline.py
─────────────────────────────────────────────────────────────────────────────
Fetches the ANBIMA Boletim de Fundos de Investimento monthly XLSX from the
ANBIMA Strapi CMS API, extracts ETF-class metrics (AUM, net flows, fund count,
and returns) and upserts them into the `anbima_etf_class_monthly` Neon table.

Data sources (all public, no auth required):
  API  : https://data-strapi.prd.anbima.com.br/api/boletim-de-fundos-de-investimentos
  XLSX : https://data-strapi.prd.anbima.com.br/<relative_path_from_api_response>

Table grain: (reference_date, anbima_type_name, metric)
Monetary values stored in R$ milhões as published by ANBIMA.
Rentabilidade values stored in percentage points (e.g. 4.37 = 4.37 %).

Sheet → metric mapping
  Pág. 4 - PL por Classe      → pl_brl_mm            (category 'ETF', monthly)
  Pág. 8 - Cap. Líq. por Classe → captacao_liquida_ytd_brl_mm (category 'ETF', annual YTD)
  Pág. 13 - N° de Fundos       → fund_count            (category 'ETF', monthly)
  Pág. 5 - PL por Tipo         → pl_brl_mm            (types 225/226, last ~16 months)
  Pág. 9 - Cap. Líq. por Tipo  → captacao_liquida_brl_mm / _ytd / _12m (types 225/226)
  Pág.11 - Rentabilidade por Tipo → rentabilidade_pct / _ytd_pct / _12m_pct (types 225/226)
"""

import asyncio
import io
import logging
import os
import re
import sys
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Tuple
import uuid

import httpx
import openpyxl

# ── repo-local imports ────────────────────────────────────────────────────────
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", ".."))
from src.store.pg_client import get_pg_client, upsert_rows  # noqa: E402

logger = logging.getLogger(__name__)

# ── constants ─────────────────────────────────────────────────────────────────
STRAPI_BASE = "https://data-strapi.prd.anbima.com.br"
STRAPI_API_PATH = (
    "/api/boletim-de-fundos-de-investimentos"
    "?populate%5Btemplate%5D%5Bpopulate%5D=attachment"
    "&pagination%5BpageSize%5D=1&sort=publishedAt%3Adesc"
)
TABLE = "anbima_etf_class_monthly"

# ANBIMA type IDs for ETF sub-categories
TYPE_ETF_RF = 225   # ETF Renda Fixa
TYPE_ETF_RV = 226   # ETF Renda Variável

# Category-level label (no type_id) — ANBIMA taxonomy uses "ETF" for the class
CATEGORY_LABEL = "ETF"

# Column indices in class-level sheets (0-indexed)
# Pág. 4 - PL por Classe  (header row 6, data from row 7)
PL_CLASSE_COL_PERIOD = 0
PL_CLASSE_COL_ETF    = 6   # "ETF" column

# Pág. 8 - Cap. Líq. por Classe  (header row 6, data from row 7)
CAPLIQ_CLASSE_COL_PERIOD = 1   # year integer
CAPLIQ_CLASSE_COL_ETF    = 7   # "ETF" column

# Pág. 13 - N° de Fundos  (header row 6, data from row 7)
NFUNDOS_COL_PERIOD = 0
NFUNDOS_COL_ETF    = 6   # "ETF" column

# Row indices in type-level sheets (0-indexed within the sheet)
# Rows 73..75 carry ETF data in Pág.5 and Pág.9:
#   row 73 = ETF aggregate (type_id=None)
#   row 74 = type_id 225 "ETF Renda Fixa"
#   row 75 = type_id 226 "ETF Renda Variável"
# Pág.11 (Rentabilidade) is handled by label scanning — no fixed row indices.
ETF_TYPE_ROWS = [73, 74, 75]   # 0-indexed — row 73=ETF agg, 74=RF(225), 75=RV(226)

# In type-level sheets:
#   col 0 = type_id (int or None), col 1 = label
#   cols 2..16 = monthly values; header row 5 has datetime objects for these cols
#   col 17 = current month value (header is a string 'abr-26', not a datetime)
#   col 18 = rolling 13-month total (Pág.9 only)
#   col 19 = YTD (Pág.9 only)
#   col 20 = 12-month total (Pág.9 only)
TYPE_COL_LABEL      = 1
TYPE_HEADER_ROW     = 5   # 0-indexed row with date objects
TYPE_COL_DATE_START = 2   # first monthly col
TYPE_COL_DATE_END   = 17  # last monthly col (header is string 'abr-26')
TYPE_COL_ROLLING    = 18  # rolling total (Pág.9 only; skipped)
TYPE_COL_YTD        = 19  # YTD total (Pág.9 only)
TYPE_COL_12M        = 20  # 12-month total (Pág.9 only)

# Map abbreviated month names (Portuguese) to month numbers
_PT_MONTH = {
    "jan": 1, "fev": 2, "mar": 3, "abr": 4, "mai": 5, "jun": 6,
    "jul": 7, "ago": 8, "set": 9, "out": 10, "nov": 11, "dez": 12,
}


# ── helpers ───────────────────────────────────────────────────────────────────

def _safe_float(v: Any) -> Optional[float]:
    """Return float or None for blank/non-numeric cells."""
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _to_date(v: Any) -> Optional[date]:
    """Convert openpyxl cell value to a date (first of the month)."""
    if isinstance(v, datetime):
        return v.date().replace(day=1)
    if isinstance(v, date):
        return v.replace(day=1)
    return None


def _parse_current_month_str(s: str) -> Optional[date]:
    """Parse 'abr-26' → date(2026, 4, 1)."""
    if not isinstance(s, str):
        return None
    m = re.match(r"([a-z]{3})-(\d{2,4})$", s.strip().lower())
    if not m:
        return None
    month_abbr, year_part = m.group(1), m.group(2)
    month_num = _PT_MONTH.get(month_abbr)
    if month_num is None:
        return None
    year = int(year_part)
    if year < 100:
        year += 2000
    return date(year, month_num, 1)


def _sheet(wb: openpyxl.Workbook, name_fragment: str) -> openpyxl.worksheet.worksheet.Worksheet:
    """Return the first sheet whose name contains name_fragment (case-insensitive)."""
    frag = name_fragment.lower()
    for sn in wb.sheetnames:
        if frag in sn.lower():
            return wb[sn]
    raise KeyError(f"Sheet containing '{name_fragment}' not found. Available: {wb.sheetnames}")


def _rows_as_lists(ws, min_row: int) -> List[List[Any]]:
    """Return all rows from min_row (1-indexed) as lists of cell values."""
    return [
        [cell.value for cell in row]
        for row in ws.iter_rows(min_row=min_row)
    ]


# ── parsers ───────────────────────────────────────────────────────────────────

def parse_pl_classe(wb: openpyxl.Workbook, boletim_ref: str) -> List[Dict]:
    """
    Pág. 4 - PL por Classe → pl_brl_mm for category 'ETF', full monthly history.
    """
    ws = _sheet(wb, "Pág. 4")
    records = []
    for row in _rows_as_lists(ws, min_row=8):   # data from row 8 (1-indexed)
        ref = _to_date(row[PL_CLASSE_COL_PERIOD])
        if ref is None:
            continue
        val = _safe_float(row[PL_CLASSE_COL_ETF])
        if val is None:
            continue
        records.append({
            "reference_date":   ref,
            "anbima_category":  CATEGORY_LABEL,
            "anbima_type_id":   None,
            "anbima_type_name": CATEGORY_LABEL,
            "metric":           "pl_brl_mm",
            "value":            val,
            "source_sheet":     "Pág. 4 - PL por Classe",
            "boletim_ref":      boletim_ref,
        })
    return records


def parse_capliq_classe(wb: openpyxl.Workbook, boletim_ref: str) -> List[Dict]:
    """
    Pág. 8 - Cap. Líq. por Classe → captacao_liquida_ytd_brl_mm for category 'ETF'.
    Annual YTD: one row per year; reference_date = last day of that year set to Dec-01
    (or current year set to Jan-01 of that year for the in-progress year).
    """
    ws = _sheet(wb, "Pág. 8")
    records = []
    for row in _rows_as_lists(ws, min_row=8):
        year_val = row[CAPLIQ_CLASSE_COL_PERIOD]
        if year_val is None:
            continue
        try:
            year = int(year_val)
        except (TypeError, ValueError):
            continue
        val = _safe_float(row[CAPLIQ_CLASSE_COL_ETF])
        if val is None:
            continue
        # Use Dec-01 for completed years; Jan-01 for current year (in-progress)
        current_year = datetime.utcnow().year
        month = 12 if year < current_year else 1
        ref = date(year, month, 1)
        records.append({
            "reference_date":   ref,
            "anbima_category":  CATEGORY_LABEL,
            "anbima_type_id":   None,
            "anbima_type_name": CATEGORY_LABEL,
            "metric":           "captacao_liquida_ytd_brl_mm",
            "value":            val,
            "source_sheet":     "Pág. 8 - Cap. Líq. por Classe",
            "boletim_ref":      boletim_ref,
        })
    return records


def parse_nfundos(wb: openpyxl.Workbook, boletim_ref: str) -> List[Dict]:
    """
    Pág. 13 - N° de Fundos → fund_count for category 'ETF', full monthly history.
    """
    ws = _sheet(wb, "Pág. 13")
    records = []
    for row in _rows_as_lists(ws, min_row=8):
        ref = _to_date(row[NFUNDOS_COL_PERIOD])
        if ref is None:
            continue
        val = _safe_float(row[NFUNDOS_COL_ETF])
        if val is None:
            continue
        records.append({
            "reference_date":   ref,
            "anbima_category":  CATEGORY_LABEL,
            "anbima_type_id":   None,
            "anbima_type_name": CATEGORY_LABEL,
            "metric":           "fund_count",
            "value":            val,
            "source_sheet":     "Pág. 13 - N° de Fundos",
            "boletim_ref":      boletim_ref,
        })
    return records


def _parse_type_sheet(
    wb: openpyxl.Workbook,
    sheet_fragment: str,
    source_sheet_label: str,
    monthly_metric: str,
    ytd_metric: str,
    twelvem_metric: str,
    boletim_ref: str,
) -> List[Dict]:
    """
    Generic parser for type-level sheets Pág.5 and Pág.9.

    Structure (April 2026 boletim):
      Header row index 5 (0-indexed) contains datetime objects for cols 2-16
      and a string 'abr-26' in col 17.
      ETF rows are at fixed offsets:
        row 73 → aggregate ETF  (type_id=None)
        row 74 → type 225 ETF Renda Fixa
        row 75 → type 226 ETF Renda Variável
      For Pág.9: col 18=rolling, col 19=YTD, col 20=12m.
      For Pág.5: cols 18-20 are None.
    """
    ws = _sheet(wb, sheet_fragment)
    all_rows = [
        [cell.value for cell in row]
        for row in ws.iter_rows()
    ]

    # Header row: index TYPE_HEADER_ROW (5)
    header_row = all_rows[TYPE_HEADER_ROW] if len(all_rows) > TYPE_HEADER_ROW else []

    # Resolve current-month date from the string in header col 17
    current_month_date: Optional[date] = None
    if len(header_row) > TYPE_COL_DATE_END:
        current_month_date = _parse_current_month_str(header_row[TYPE_COL_DATE_END])
    # Fallback: advance from the last datetime in header
    if current_month_date is None:
        for ci in range(TYPE_COL_DATE_END - 1, TYPE_COL_DATE_START - 1, -1):
            d = _to_date(header_row[ci]) if ci < len(header_row) else None
            if d:
                current_month_date = (
                    date(d.year + 1, 1, 1) if d.month == 12
                    else date(d.year, d.month + 1, 1)
                )
                break

    type_id_map: Dict[int, Optional[int]] = {
        ETF_TYPE_ROWS[0]: None,
        ETF_TYPE_ROWS[1]: TYPE_ETF_RF,
        ETF_TYPE_ROWS[2]: TYPE_ETF_RV,
    }
    type_name_map: Dict[int, str] = {
        ETF_TYPE_ROWS[0]: CATEGORY_LABEL,
        ETF_TYPE_ROWS[1]: "ETF Renda Fixa",
        ETF_TYPE_ROWS[2]: "ETF Renda Variável",
    }

    records = []

    for row_idx in ETF_TYPE_ROWS:
        if row_idx >= len(all_rows):
            logger.warning(
                "Sheet '%s': expected row %d not present (sheet has %d rows)",
                sheet_fragment, row_idx, len(all_rows)
            )
            continue
        row = all_rows[row_idx]
        type_id   = type_id_map[row_idx]
        type_name = type_name_map[row_idx]

        # ── monthly cols 2..16 (header has datetime objects) ──────────────
        for col_idx in range(TYPE_COL_DATE_START, TYPE_COL_DATE_END):  # 2..16
            if col_idx >= len(row):
                break
            val = _safe_float(row[col_idx])
            if val is None:
                continue
            ref = _to_date(header_row[col_idx]) if col_idx < len(header_row) else None
            if ref is None:
                continue
            records.append({
                "reference_date":   ref,
                "anbima_category":  CATEGORY_LABEL,
                "anbima_type_id":   type_id,
                "anbima_type_name": type_name,
                "metric":           monthly_metric,
                "value":            val,
                "source_sheet":     source_sheet_label,
                "boletim_ref":      boletim_ref,
            })

        # ── current month col 17 (header is a string, not datetime) ───────
        if TYPE_COL_DATE_END < len(row) and current_month_date:
            cur_val = _safe_float(row[TYPE_COL_DATE_END])
            if cur_val is not None:
                records.append({
                    "reference_date":   current_month_date,
                    "anbima_category":  CATEGORY_LABEL,
                    "anbima_type_id":   type_id,
                    "anbima_type_name": type_name,
                    "metric":           monthly_metric,
                    "value":            cur_val,
                    "source_sheet":     source_sheet_label,
                    "boletim_ref":      boletim_ref,
                })

        anchor = current_month_date

        # ── YTD total (col 19) ────────────────────────────────────────────
        if ytd_metric and TYPE_COL_YTD < len(row) and anchor:
            ytd_val = _safe_float(row[TYPE_COL_YTD])
            if ytd_val is not None:
                records.append({
                    "reference_date":   anchor,
                    "anbima_category":  CATEGORY_LABEL,
                    "anbima_type_id":   type_id,
                    "anbima_type_name": type_name,
                    "metric":           ytd_metric,
                    "value":            ytd_val,
                    "source_sheet":     source_sheet_label,
                    "boletim_ref":      boletim_ref,
                })

        # ── 12-month total (col 20) ───────────────────────────────────────
        if twelvem_metric and TYPE_COL_12M < len(row) and anchor:
            twm_val = _safe_float(row[TYPE_COL_12M])
            if twm_val is not None:
                records.append({
                    "reference_date":   anchor,
                    "anbima_category":  CATEGORY_LABEL,
                    "anbima_type_id":   type_id,
                    "anbima_type_name": type_name,
                    "metric":           twelvem_metric,
                    "value":            twm_val,
                    "source_sheet":     source_sheet_label,
                    "boletim_ref":      boletim_ref,
                })

    return records

def parse_pl_tipo(wb: openpyxl.Workbook, boletim_ref: str) -> List[Dict]:
    """Pág. 5 - PL por Tipo → pl_brl_mm for ETF types (last ~16 months)."""
    return _parse_type_sheet(
        wb, "Pág. 5", "Pág. 5 - PL por Tipo",
        monthly_metric="pl_brl_mm",
        ytd_metric="",      # no YTD on PL sheet
        twelvem_metric="",
        boletim_ref=boletim_ref,
    )


def parse_capliq_tipo(wb: openpyxl.Workbook, boletim_ref: str) -> List[Dict]:
    """Pág. 9 - Cap. Líq. por Tipo → captacao_liquida_* for ETF types."""
    return _parse_type_sheet(
        wb, "Pág. 9", "Pág. 9 - Cap. Líq. por Tipo",
        monthly_metric="captacao_liquida_brl_mm",
        ytd_metric="captacao_liquida_ytd_brl_mm",
        twelvem_metric="captacao_liquida_12m_brl_mm",
        boletim_ref=boletim_ref,
    )


def parse_rentabilidade(wb: openpyxl.Workbook, boletim_ref: str) -> List[Dict]:
    """
    Pág.11 - Rentabilidade por Tipo → rentabilidade_pct / _ytd_pct / _12m_pct.

    NOTE: ANBIMA does not always include ETF rows in this sheet.
    We scan every row by label to be robust across boletim editions.

    Matched labels (case-insensitive, strip-prefix):
      starts with 'etf renda fix' → type 225 (ETF Renda Fixa)
      starts with 'etf renda var' → type 226 (ETF Renda Variável)
      exactly 'etf'               → aggregate (type_id=None)
    """
    ws = _sheet(wb, "Pág.11")
    all_rows = [
        [cell.value for cell in row]
        for row in ws.iter_rows()
    ]

    # Find header row: first row where col2 is a datetime object
    header_row: List[Any] = []
    for r in all_rows:
        if len(r) > 2 and isinstance(r[2], datetime):
            header_row = r
            break
    if not header_row:
        logger.info(
            "Pág.11: no header row with date objects found — "
            "ANBIMA may not publish ETF returns in this boletim edition."
        )
        return []

    # Resolve current-month date
    current_month_date: Optional[date] = None
    if len(header_row) > TYPE_COL_DATE_END:
        current_month_date = _parse_current_month_str(header_row[TYPE_COL_DATE_END])
    if current_month_date is None:
        for ci in range(TYPE_COL_DATE_END - 1, TYPE_COL_DATE_START - 1, -1):
            d = _to_date(header_row[ci]) if ci < len(header_row) else None
            if d:
                current_month_date = (
                    date(d.year + 1, 1, 1) if d.month == 12
                    else date(d.year, d.month + 1, 1)
                )
                break

    def _match_etf_label(label: Any) -> Optional[tuple]:
        """Return (type_id, type_name) for ETF rows, None otherwise."""
        if not isinstance(label, str):
            return None
        lab = label.strip().lower()
        if lab.startswith("etf renda fix"):
            return (TYPE_ETF_RF, "ETF Renda Fixa")
        if lab.startswith("etf renda var"):
            return (TYPE_ETF_RV, "ETF Renda Variável")
        if lab == "etf":
            return (None, CATEGORY_LABEL)
        return None

    records = []
    source_label = "Pág.11 - Rentabilidade por Tipo"

    for row in all_rows:
        if not row or len(row) < 3:
            continue
        info = _match_etf_label(row[1] if len(row) > 1 else None)
        if info is None:
            continue
        type_id, type_name = info

        # Monthly cols 2..16
        for col_idx in range(TYPE_COL_DATE_START, TYPE_COL_DATE_END):
            if col_idx >= len(row):
                break
            val = _safe_float(row[col_idx])
            if val is None:
                continue
            ref = _to_date(header_row[col_idx]) if col_idx < len(header_row) else None
            if ref is None:
                continue
            records.append({
                "reference_date":   ref,
                "anbima_category":  CATEGORY_LABEL,
                "anbima_type_id":   type_id,
                "anbima_type_name": type_name,
                "metric":           "rentabilidade_pct",
                "value":            val,
                "source_sheet":     source_label,
                "boletim_ref":      boletim_ref,
            })

        # Current month col 17
        anchor = current_month_date
        if TYPE_COL_DATE_END < len(row) and anchor:
            cur_val = _safe_float(row[TYPE_COL_DATE_END])
            if cur_val is not None:
                records.append({
                    "reference_date":   anchor,
                    "anbima_category":  CATEGORY_LABEL,
                    "anbima_type_id":   type_id,
                    "anbima_type_name": type_name,
                    "metric":           "rentabilidade_pct",
                    "value":            cur_val,
                    "source_sheet":     source_label,
                    "boletim_ref":      boletim_ref,
                })

        if anchor:
            # YTD (col 19)
            if TYPE_COL_YTD < len(row):
                ytd_val = _safe_float(row[TYPE_COL_YTD])
                if ytd_val is not None:
                    records.append({
                        "reference_date":   anchor,
                        "anbima_category":  CATEGORY_LABEL,
                        "anbima_type_id":   type_id,
                        "anbima_type_name": type_name,
                        "metric":           "rentabilidade_ytd_pct",
                        "value":            ytd_val,
                        "source_sheet":     source_label,
                        "boletim_ref":      boletim_ref,
                    })
            # 12m (col 20)
            if TYPE_COL_12M < len(row):
                twm_val = _safe_float(row[TYPE_COL_12M])
                if twm_val is not None:
                    records.append({
                        "reference_date":   anchor,
                        "anbima_category":  CATEGORY_LABEL,
                        "anbima_type_id":   type_id,
                        "anbima_type_name": type_name,
                        "metric":           "rentabilidade_12m_pct",
                        "value":            twm_val,
                        "source_sheet":     source_label,
                        "boletim_ref":      boletim_ref,
                    })

    if not records:
        logger.info(
            "Pág.11: no ETF rentabilidade rows found — "
            "ANBIMA does not include ETF returns in this boletim edition (expected for some months)."
        )
    return records



# ── ANBIMA API ────────────────────────────────────────────────────────────────

async def fetch_latest_boletim_url() -> Tuple[str, str]:
    """
    Query the ANBIMA Strapi API and return (download_url, boletim_ref).
    boletim_ref = the upload filename slug (used for provenance).
    """
    url = STRAPI_BASE + STRAPI_API_PATH
    async with httpx.AsyncClient(timeout=30) as client:
        resp = await client.get(url)
        resp.raise_for_status()
    data = resp.json()
    try:
        rel_path = (
            data["data"][0]["attributes"]["template"]["attachment"]["data"]["attributes"]["url"]
        )
    except (KeyError, IndexError, TypeError) as exc:
        raise RuntimeError(f"Unexpected ANBIMA Strapi response structure: {exc}") from exc
    download_url = STRAPI_BASE + rel_path
    boletim_ref  = rel_path.split("/")[-1]   # e.g. "Anexo_boletim_fundos_investimento_abril_valor_7e6dc82403.xlsx"
    return download_url, boletim_ref


async def download_xlsx(url: str) -> bytes:
    """Download the XLSX file and return raw bytes."""
    async with httpx.AsyncClient(timeout=120, follow_redirects=True) as client:
        resp = await client.get(url)
        resp.raise_for_status()
    return resp.content


def parse_boletim(xlsx_bytes: bytes, boletim_ref: str) -> List[Dict]:
    """
    Parse all relevant sheets from the XLSX bytes and return a flat list of
    record dicts ready for upsert into anbima_etf_class_monthly.
    """
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)

    records: List[Dict] = []
    parsers = [
        ("Pág. 4 (PL classe)",      parse_pl_classe),
        ("Pág. 8 (CapLiq classe)",  parse_capliq_classe),
        ("Pág. 13 (N° fundos)",     parse_nfundos),
        ("Pág. 5 (PL tipo)",        parse_pl_tipo),
        ("Pág. 9 (CapLiq tipo)",    parse_capliq_tipo),
        ("Pág. 11 (Rentabilidade)", parse_rentabilidade),
    ]
    for label, fn in parsers:
        try:
            batch = fn(wb, boletim_ref)
            logger.info("  %s → %d records", label, len(batch))
            records.extend(batch)
        except Exception as exc:
            logger.warning("Parser failed for %s: %s", label, exc)

    wb.close()
    return records


# ── ingestor class ────────────────────────────────────────────────────────────

class AnbimaEtfIngestor:
    """
    Fetches and upserts ANBIMA ETF class metrics.

    Public interface (mirrors other pipeline classes in this repo):
      daily_update()  — fetch latest boletim and upsert; idempotent
      backfill()      — not applicable here (historical data is embedded in
                        every boletim XLSX); daily_update() already covers
                        full history from each file.
    """

    def __init__(self):
        self._pg = get_pg_client()

    # ── internal helpers ──────────────────────────────────────────────────────

    def _log_start(self, conn, run_id: str, boletim_ref: str) -> None:
        upsert_rows(
            conn,
            "cvm_ingest_log",  # reuse shared log table (matches cvm_pipeline pattern)
            [{
                "run_id":       run_id,
                "entity":       "anbima_etf",
                "doc_type":     "boletim_mensal",
                "period_year":  None,
                "period_month": None,
                "status":       "running",
                "started_at":   datetime.utcnow(),
                "notes":        boletim_ref,
            }],
            conflict_columns="run_id",
        )

    def _log_finish(
        self,
        conn,
        run_id: str,
        status: str,
        rows_upserted: int,
        error: Optional[str] = None,
    ) -> None:
        upsert_rows(
            conn,
            "cvm_ingest_log",
            [{
                "run_id":         run_id,
                "entity":         "anbima_etf",
                "doc_type":       "boletim_mensal",
                "status":         status,
                "rows_upserted":  rows_upserted,
                "finished_at":    datetime.utcnow(),
                "error_message":  error,
            }],
            conflict_columns="run_id",
        )

    # ── public API ────────────────────────────────────────────────────────────

    async def daily_update(self) -> Dict[str, int]:
        """
        Fetch the latest published ANBIMA boletim and upsert all ETF metrics.
        Idempotent: re-running with the same boletim inserts/updates the same rows.

        Returns: {'anbima_etf': <rows_upserted>}
        """
        run_id = str(uuid.uuid4())
        rows_upserted = 0

        logger.info("[anbima_etf] Starting daily update (run_id=%s)", run_id)
        download_url, boletim_ref = await fetch_latest_boletim_url()
        logger.info("[anbima_etf] Latest boletim: %s", boletim_ref)

        xlsx_bytes = await download_xlsx(download_url)
        logger.info("[anbima_etf] Downloaded %d bytes", len(xlsx_bytes))

        records = parse_boletim(xlsx_bytes, boletim_ref)
        logger.info("[anbima_etf] Parsed %d records total", len(records))

        if not records:
            logger.warning("[anbima_etf] No records parsed — skipping upsert")
            return {"anbima_etf": 0}

        self._log_start(self._pg, run_id, boletim_ref)
        try:
            upsert_rows(
                self._pg,
                TABLE,
                records,
                conflict_columns="reference_date,anbima_type_name,metric",
            )
            rows_upserted = len(records)
            self._log_finish(self._pg, run_id, "success", rows_upserted)
            logger.info("[anbima_etf] Upserted %d rows", rows_upserted)
        except Exception as exc:
            self._log_finish(self._pg, run_id, "error", 0, str(exc))
            logger.error("[anbima_etf] Upsert failed: %s", exc)
            raise

        return {"anbima_etf": rows_upserted}

    async def backfill(self, start_year: int = 2006) -> Dict[str, int]:
        """
        The ANBIMA boletim XLSX already contains full history back to 2006 in
        the class-level sheets (Pág.4, Pág.8, Pág.13). Calling daily_update()
        is sufficient to backfill; this method is a convenience alias.
        """
        logger.info(
            "[anbima_etf] backfill() called — delegating to daily_update() "
            "(boletim already contains full history back to %d)", start_year
        )
        return await self.daily_update()


# ── CLI entry point ───────────────────────────────────────────────────────────

if __name__ == "__main__":
    import argparse

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)-8s %(name)s: %(message)s",
    )

    parser = argparse.ArgumentParser(description="ANBIMA ETF boletim ingestor")
    parser.add_argument(
        "--mode",
        choices=["daily", "backfill"],
        default="daily",
        help="daily: fetch latest boletim; backfill: alias for daily (full history embedded)",
    )
    args = parser.parse_args()

    ingestor = AnbimaEtfIngestor()
    if args.mode == "backfill":
        result = asyncio.run(ingestor.backfill())
    else:
        result = asyncio.run(ingestor.daily_update())
    print(result)
