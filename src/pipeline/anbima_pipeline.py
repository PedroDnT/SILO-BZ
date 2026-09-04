"""
anbima_pipeline.py
─────────────────────────────────────────────────────────────────────────────
Fetches the ANBIMA "Boletim de Fundos de Investimento" monthly XLSX from the
ANBIMA Strapi CMS API and upserts **every ANBIMA class and type** it publishes
into the `anbima_class_monthly` Supabase table.

Data sources (all public, no auth required):
  API  : https://data-strapi.prd.anbima.com.br/api/boletim-de-fundos-de-investimentos
  XLSX : https://data-strapi.prd.anbima.com.br/<relative_path_from_api_response>

Table grain: (reference_date, anbima_category, anbima_type_name, metric, level)
Monetary values stored in R$ milhões as published by ANBIMA.
Rentabilidade values stored in percentage points (e.g. 4.37 = 4.37 %).

Hierarchy — the `level` column
──────────────────────────────
The boletim's type sheets are hierarchical: a class-aggregate row, then the
ANBIMA types under it, then industry totals. `level` records which of those a row
came from, and it is part of the key because the labels **Cambial**, **FIP** and
**FIAGRO** each appear TWICE in the same sheet — once as the class aggregate and
once as an ANBIMA type of the identical name (ids 251 / 238 / 348). Without
`level` (and the owning category) the type row silently overwrites the class
aggregate.

  'category' → class aggregate  (anbima_type_id is NULL, anbima_category = itself)
  'type'     → ANBIMA type      (anbima_type_id set when the sheet publishes one)
  'total'    → industry total   (anbima_category = 'TOTAL'; belongs to no class)

Sheet → metric mapping
  Pág. 4 - PL por Classe        → pl_brl_mm                    (per class, full monthly history)
  Pág. 8 - Cap. Líq. por Classe → captacao_liquida_brl_mm      (monthly rows, YYYYMM in col 0)
                                  captacao_liquida_ytd_brl_mm  (annual rows,  YYYY   in col 0)
  Pág. 13 - N° de Fundos        → fund_count                   (per class, full monthly history)
  Pág. 5 - PL por Tipo          → pl_brl_mm                    (classes + types, ~19-month window)
  Pág. 9 - Cap. Líq. por Tipo   → captacao_liquida_brl_mm / _ytd_ / _12m_
  Pág.11 - Rentabilidade por Tipo → rentabilidade_pct / _ytd_pct / _12m_pct

Nothing here is positional. Header rows, data ranges and value columns are all
DISCOVERED from the sheet (dates in the header, class names in the header, type
ids in column 0), because ANBIMA reflows these workbooks between editions and a
hardcoded offset does not fail loudly — it keeps reading *some other category's*
numbers and stores them under the wrong label. A row we cannot interpret is
skipped and counted, never guessed at.

Layout facts this replaces (all verified against the July-2026 boletim):
  • class sheets used to be read from a hardcoded row, which silently dropped
    the first published year (2006) from every class series;
  • Pág. 8 mixes annual rows (col 0 = 2025) with monthly rows (col 0 = 202501);
    the old reader treated the monthly rows' *month number* as a year and
    anchored them at year 0001;
  • the type sheets grew from 16 to 19 month columns, so the old fixed YTD /
    12-month column offsets had drifted onto plain monthly columns.
"""

import asyncio
import io
import logging
import os
import re
import sys
import unicodedata
from datetime import date, datetime, timezone
from typing import Any, Dict, List, Optional, Sequence, Tuple
import uuid

import httpx
import openpyxl

# ── repo-local imports ────────────────────────────────────────────────────────
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", ".."))
from src.store.pg_client import get_pg_client, upsert_rows  # noqa: E402
from src.pipeline import ingest_log  # noqa: E402

logger = logging.getLogger(__name__)

# ── constants ─────────────────────────────────────────────────────────────────
STRAPI_BASE = "https://data-strapi.prd.anbima.com.br"
STRAPI_API_PATH = (
    "/api/boletim-de-fundos-de-investimentos"
    "?populate%5Btemplate%5D%5Bpopulate%5D=attachment"
    "&pagination%5BpageSize%5D=1&sort=publishedAt%3Adesc"
)
TABLE = "anbima_class_monthly"
CONFLICT_COLUMNS = "reference_date,anbima_category,anbima_type_name,metric,level"

# Audit-log identity. Deliberately unchanged from the ETF-only era:
# scripts/check_staleness.py and the existing cvm_ingest_log history key on it,
# and renaming the entity would orphan both.
LOG_ENTITY = "anbima_etf"
LOG_DOC_TYPE = "boletim_mensal"

LEVEL_CATEGORY = "category"
LEVEL_TYPE = "type"
LEVEL_TOTAL = "total"

# Category assigned to industry-total rows. A total spans every class, so it
# belongs to none of them; anbima_category is NOT NULL, hence an explicit label.
TOTAL_CATEGORY = "TOTAL"

# The 11 ANBIMA classes, in the canonical spelling this pipeline stores. ANBIMA
# writes the same class differently across sheets ('Renda Fixa' / 'Renda fixa',
# 'OFF-SHORE' / 'Off shore') and decorates some with footnote markers
# ('FIAGRO (11)'). Keys are the accent/case/punctuation-stripped form so every
# spelling lands on ONE key — without that the same series would split into
# several primary keys across sheets and across editions.
CANONICAL_CATEGORIES: Dict[str, str] = {
    "rendafixa":     "Renda Fixa",
    "acoes":         "Ações",
    "multimercados": "Multimercados",
    "cambial":       "Cambial",
    "previdencia":   "Previdência",
    "etf":           "ETF",
    "fidc":          "FIDC",
    "fip":           "FIP",
    "fiagro":        "FIAGRO",
    "fii":           "FII",
    "offshore":      "Off Shore",
}

# Industry totals, same normalisation problem: Pág. 5 says 'Total Fundos de
# Investimentos' where Pág. 9 says 'Total fundos de investimento'.
CANONICAL_TOTALS: Dict[str, str] = {
    "totalfundosdeinvestimento":  "Total Fundos de Investimento",
    "totalfundosdeinvestimentos": "Total Fundos de Investimento",
    "totalfundosestruturados":    "Total Fundos Estruturados",
    "totaldomestico":             "Total Doméstico",
    "totalfundosoffshore":        "Total Fundos Off Shore",
    "totalgeral":                 "Total Geral",
}

# Metric names
METRIC_PL = "pl_brl_mm"
METRIC_CAPLIQ = "captacao_liquida_brl_mm"
METRIC_CAPLIQ_YTD = "captacao_liquida_ytd_brl_mm"
METRIC_CAPLIQ_12M = "captacao_liquida_12m_brl_mm"
METRIC_FUND_COUNT = "fund_count"
METRIC_RENT = "rentabilidade_pct"
METRIC_RENT_YTD = "rentabilidade_ytd_pct"
METRIC_RENT_12M = "rentabilidade_12m_pct"

# In every type sheet: col 0 = ANBIMA type id, col 1 = label, col 2+ = values.
TYPE_COL_ID = 0
TYPE_COL_LABEL = 1
TYPE_COL_FIRST_VALUE = 2

# Map abbreviated month names (Portuguese) to month numbers
_PT_MONTH = {
    "jan": 1, "fev": 2, "mar": 3, "abr": 4, "mai": 5, "jun": 6,
    "jul": 7, "ago": 8, "set": 9, "out": 10, "nov": 11, "dez": 12,
}

# Footnote markers ANBIMA appends to labels: 'FIAGRO (11)', 'Tipos ANBIMA(10)'.
_FOOTNOTE_RE = re.compile(r"\s*\(\d+\)\s*$")


# ── text helpers ──────────────────────────────────────────────────────────────

def strip_footnote(label: Any) -> Optional[str]:
    """'FIAGRO (11)' → 'FIAGRO'; collapse whitespace; None for non-strings.

    Footnote markers move between editions (a class picks one up, another drops
    it). Leaving them in the name would fork the primary key on the next boletim,
    so they are stripped before the value is keyed.
    """
    if not isinstance(label, str):
        return None
    cleaned = _FOOTNOTE_RE.sub("", label)
    cleaned = " ".join(cleaned.split())
    return cleaned or None


def _key(label: str) -> str:
    """Accent/case/punctuation-insensitive lookup key ('Off shore' → 'offshore')."""
    decomposed = unicodedata.normalize("NFKD", label)
    ascii_only = "".join(c for c in decomposed if not unicodedata.combining(c))
    return "".join(c for c in ascii_only.lower() if c.isalnum())


def _ascii_lower(s: str) -> str:
    """Lowercase, accents removed, whitespace preserved (for header matching)."""
    decomposed = unicodedata.normalize("NFKD", s)
    return "".join(c for c in decomposed if not unicodedata.combining(c)).lower()


def canonical_category(label: Any) -> Optional[str]:
    """Return the canonical ANBIMA class name for a label, else None."""
    cleaned = strip_footnote(label)
    if cleaned is None:
        return None
    return CANONICAL_CATEGORIES.get(_key(cleaned))


def is_total_label(label: Any) -> bool:
    """True for the industry-total rows ('Total geral', 'Total doméstico', …)."""
    cleaned = strip_footnote(label)
    if cleaned is None:
        return False
    return _ascii_lower(cleaned).startswith("total")


def canonical_total(label: Any) -> Optional[str]:
    """Canonical name for an industry-total row; unknown totals keep their label."""
    cleaned = strip_footnote(label)
    if cleaned is None:
        return None
    return CANONICAL_TOTALS.get(_key(cleaned), cleaned)


# ── cell helpers ──────────────────────────────────────────────────────────────

def _safe_float(v: Any) -> Optional[float]:
    """Return a finite float, or None for blank / non-numeric / NaN / inf cells.

    NaN and inf reach us from Excel error cells (#DIV/0!, #N/A). They are not
    data; storing them would poison every downstream aggregate.
    """
    if v is None or isinstance(v, bool):
        return None
    try:
        f = float(v)
    except (TypeError, ValueError):
        return None
    if f != f or f in (float("inf"), float("-inf")):
        return None
    return f


def _to_date(v: Any) -> Optional[date]:
    """Convert an openpyxl cell value to a date (first of the month)."""
    if isinstance(v, datetime):
        return v.date().replace(day=1)
    if isinstance(v, date):
        return v.replace(day=1)
    return None


def _parse_current_month_str(s: Any) -> Optional[date]:
    """Parse 'abr-26' → date(2026, 4, 1)."""
    if not isinstance(s, str):
        return None
    m = re.match(r"([a-z]{3})-(\d{2,4})$", _ascii_lower(s).strip())
    if not m:
        return None
    month_num = _PT_MONTH.get(m.group(1))
    if month_num is None:
        return None
    year = int(m.group(2))
    if year < 100:
        year += 2000
    return date(year, month_num, 1)


def _as_int(v: Any) -> Optional[int]:
    """Return an int for integral numeric cells ('225', 225, 225.0), else None."""
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        return int(v) if v.is_integer() else None
    if isinstance(v, str):
        s = v.strip()
        return int(s) if s.isdigit() else None
    return None


def _sheet(wb: openpyxl.Workbook, name_fragment: str):
    """Return the first sheet whose name contains name_fragment (case-insensitive)."""
    frag = name_fragment.lower()
    for sn in wb.sheetnames:
        if frag in sn.lower():
            return wb[sn]
    raise KeyError(f"Sheet containing '{name_fragment}' not found. Available: {wb.sheetnames}")


def _rows(ws) -> List[List[Any]]:
    return [[cell.value for cell in row] for row in ws.iter_rows()]


def _record(
    reference_date: date,
    category: str,
    type_id: Optional[int],
    type_name: str,
    metric: str,
    value: float,
    level: str,
    source_sheet: str,
    boletim_ref: str,
) -> Dict[str, Any]:
    return {
        "reference_date":   reference_date,
        "anbima_category":  category,
        "anbima_type_id":   type_id,
        "anbima_type_name": type_name,
        "metric":           metric,
        "value":            value,
        "level":            level,
        "source_sheet":     source_sheet,
        "boletim_ref":      boletim_ref,
    }


# ── class sheets (Pág. 4 / 8 / 13) — one column per ANBIMA class ──────────────

def _class_period(row: Sequence[Any]) -> Optional[Tuple[date, bool]]:
    """Interpret column 0 of a class sheet as (reference_date, is_annual_total).

    Three shapes occur, all straight from the sheet:
      datetime(2026, 7, 1) → a month            (Pág. 4 / Pág. 13)
      2026                 → a whole year       (Pág. 8 annual rows)
      202607               → a month as YYYYMM  (Pág. 8 monthly rows)
    """
    if not row:
        return None
    raw = row[0]

    as_date = _to_date(raw)
    if as_date is not None:
        return as_date, False

    n = _as_int(raw)
    if n is None:
        return None
    if 1900 <= n <= 2100:
        # Annual row. Anchor completed years at December and the in-progress
        # year at January — the convention the ETF series was built on.
        current_year = datetime.now(timezone.utc).year
        return date(n, 12 if n < current_year else 1, 1), True
    if 190001 <= n <= 210012:
        year, month = divmod(n, 100)
        if 1 <= month <= 12:
            return date(year, month, 1), False
    return None


def find_class_header(rows: List[List[Any]]) -> Tuple[int, List[Any]]:
    """Locate the class-name header row of a wide sheet.

    Returns (first_data_row_index, header_row). ANBIMA writes the class names
    twice: a clean row, then a footnoted row that may also carry 'Total' columns,
    then the data. So the clean row is two above the first data row — but the
    header is *found* by walking down to the first row whose column 0 parses as a
    period, never by a hardcoded index. Reading from a hardcoded row is what
    silently dropped the 2006 line from every series.
    """
    for idx, row in enumerate(rows):
        if _class_period(row) is None:
            continue
        header_idx = idx - 2
        if header_idx >= 0 and sum(
            1 for v in rows[header_idx] if isinstance(v, str) and v.strip()
        ) >= 2:
            return idx, rows[header_idx]
        # Fall back to the footnoted row; strip_footnote normalises it anyway.
        if idx - 1 >= 0:
            return idx, rows[idx - 1]
        return idx, []
    return -1, []


def parse_class_sheet(
    wb: openpyxl.Workbook,
    sheet_fragment: str,
    source_sheet_label: str,
    monthly_metric: str,
    annual_metric: str,
    boletim_ref: str,
) -> List[Dict]:
    """Parse a wide class sheet: one column per ANBIMA class, one row per period.

    `monthly_metric` is emitted for month rows and `annual_metric` for whole-year
    rows (Pág. 8 publishes both in one sheet). Pass "" to drop a shape.
    """
    ws = _sheet(wb, sheet_fragment)
    rows = _rows(ws)

    first_data_idx, header = find_class_header(rows)
    if first_data_idx < 0:
        logger.warning(
            "Sheet '%s': no data row found (column 0 never parses as a period) — "
            "emitting nothing rather than reading blind offsets", sheet_fragment,
        )
        return []

    # column index → canonical class name
    columns: Dict[int, str] = {}
    for col_idx, raw_name in enumerate(header):
        name = strip_footnote(raw_name)
        if not name:
            continue
        category = canonical_category(name)
        if category is not None:
            columns[col_idx] = category
        elif not is_total_label(name):
            # Not a class and not one of the sheet's Total columns.
            logger.warning(
                "Sheet '%s': header column %d %r is not a known ANBIMA class — "
                "skipping the column", sheet_fragment, col_idx, name,
            )
    if not columns:
        logger.warning("Sheet '%s': no ANBIMA class columns in the header row",
                       sheet_fragment)
        return []

    records: List[Dict] = []
    skipped = 0
    for row in rows[first_data_idx:]:
        period = _class_period(row)
        if period is None:
            skipped += 1
            continue
        ref, is_annual = period
        metric = annual_metric if is_annual else monthly_metric
        if not metric:
            continue
        for col_idx, category in columns.items():
            if col_idx >= len(row):
                continue
            val = _safe_float(row[col_idx])
            if val is None:
                continue
            records.append(_record(
                ref, category, None, category, metric, val,
                LEVEL_CATEGORY, source_sheet_label, boletim_ref,
            ))

    logger.debug("Sheet '%s': %d classes, %d records, %d non-data rows skipped",
                 sheet_fragment, len(columns), len(records), skipped)
    return records


def parse_pl_classe(wb: openpyxl.Workbook, boletim_ref: str) -> List[Dict]:
    """Pág. 4 - PL por Classe → pl_brl_mm per class, full monthly history."""
    return parse_class_sheet(
        wb, "Pág. 4", "Pág. 4 - PL por Classe",
        monthly_metric=METRIC_PL, annual_metric=METRIC_PL,
        boletim_ref=boletim_ref,
    )


def parse_capliq_classe(wb: openpyxl.Workbook, boletim_ref: str) -> List[Dict]:
    """Pág. 8 - Cap. Líq. por Classe → net flows per class.

    The sheet interleaves whole-year rows (col 0 = 2025) with the months of the
    last two years (col 0 = 202501). The annual rows are the year's accumulated
    net flow → captacao_liquida_ytd_brl_mm; the monthly rows are that single
    month's net flow → captacao_liquida_brl_mm (verified: they match Pág. 9's
    monthly columns exactly).
    """
    return parse_class_sheet(
        wb, "Pág. 8", "Pág. 8 - Cap. Líq. por Classe",
        monthly_metric=METRIC_CAPLIQ, annual_metric=METRIC_CAPLIQ_YTD,
        boletim_ref=boletim_ref,
    )


def parse_nfundos(wb: openpyxl.Workbook, boletim_ref: str) -> List[Dict]:
    """Pág. 13 - N° de Fundos → fund_count per class, full monthly history."""
    return parse_class_sheet(
        wb, "Pág. 13", "Pág. 13 - N° de Fundos",
        monthly_metric=METRIC_FUND_COUNT, annual_metric=METRIC_FUND_COUNT,
        boletim_ref=boletim_ref,
    )


# ── type sheets (Pág. 5 / 9 / 11) — long + hierarchical ──────────────────────

def find_type_header(rows: List[List[Any]]) -> int:
    """Index of the header row: the first row carrying date objects in col 2+."""
    for idx, row in enumerate(rows):
        if any(isinstance(v, datetime) for v in row[TYPE_COL_FIRST_VALUE:]):
            return idx
    return -1


def map_type_columns(header: Sequence[Any]) -> Tuple[Dict[int, date], List[int], List[int]]:
    """Classify a type sheet's value columns from its header row.

    Returns (month columns → date, YTD columns, 12-month columns). Everything
    else — notably the rolling 'jul/25 até jul/26' window, whose period is
    ambiguous — is skipped. Discovered, not hardcoded: between editions ANBIMA
    grew the monthly block from 16 to 19 columns, which slid the old fixed YTD /
    12-month offsets onto ordinary monthly values.
    """
    months: Dict[int, date] = {}
    ytd_cols: List[int] = []
    twelve_m_cols: List[int] = []

    for col_idx in range(TYPE_COL_FIRST_VALUE, len(header)):
        raw = header[col_idx]

        as_date = _to_date(raw)
        if as_date is not None:
            months[col_idx] = as_date
            continue
        if not isinstance(raw, str) or not raw.strip():
            continue

        as_month = _parse_current_month_str(raw)
        if as_month is not None:
            months[col_idx] = as_month
            continue

        label = _ascii_lower(raw)
        if "12 meses" in label:
            twelve_m_cols.append(col_idx)
        elif re.search(r"\bano\b", label):
            ytd_cols.append(col_idx)
        else:
            logger.debug("Type sheet header column %d %r not a period — skipped",
                         col_idx, raw)

    return months, ytd_cols, twelve_m_cols


def classify_type_row(
    row: Sequence[Any],
    current_category: Optional[str],
    has_values: bool,
) -> Optional[Tuple[str, Optional[int], str, str]]:
    """Decide what a type-sheet row is.

    Returns (level, type_id, type_name, category) or None when the row carries
    nothing we can attribute (section banner, repeated header, footnote, or an
    empty type row) — those are skipped and counted, never guessed at.

    Rules, in order:
      • 'Total …'          → an industry total. Never changes the current class,
                             or 'Total geral' at the foot of the sheet would
                             become the class of everything after it.
      • type id present    → an ANBIMA type of the current class.
      • label is a class   → the class aggregate; becomes the current class.
                             (Pág. 11 prints these with no values at all, so a
                             row is a class because of its NAME, not its data.)
      • label starts with  → an unnumbered type of the current class (Pág. 9
        the current class     prints the FII types with no ids).
      • anything else with → an unknown class. Better a loud new category than
        values                silently filing it under the previous one.
    """
    label = strip_footnote(row[TYPE_COL_LABEL] if len(row) > TYPE_COL_LABEL else None)
    if not label:
        return None

    if is_total_label(label):
        return (LEVEL_TOTAL, None, canonical_total(label) or label, TOTAL_CATEGORY)

    type_id = _as_int(row[TYPE_COL_ID]) if row else None
    if type_id is not None:
        if current_category is None:
            logger.warning("ANBIMA type %r (id %s) appears before any class row — "
                           "skipped rather than filed under a guessed class",
                           label, type_id)
            return None
        return (LEVEL_TYPE, type_id, label, current_category)

    category = canonical_category(label)
    if category is not None:
        return (LEVEL_CATEGORY, None, category, category)

    if not has_values:
        return None   # section banner, repeated header, footnote, empty row

    if current_category is not None and _key(label).startswith(_key(current_category)):
        return (LEVEL_TYPE, None, label, current_category)

    logger.warning("ANBIMA row %r is not a known class but carries values — "
                   "storing it as its own category", label)
    return (LEVEL_CATEGORY, None, label, label)


def parse_type_sheet(
    wb: openpyxl.Workbook,
    sheet_fragment: str,
    source_sheet_label: str,
    monthly_metric: str,
    ytd_metric: str,
    twelvem_metric: str,
    boletim_ref: str,
) -> List[Dict]:
    """Parse a long, hierarchical type sheet (Pág. 5 / Pág. 9 / Pág. 11)."""
    ws = _sheet(wb, sheet_fragment)
    rows = _rows(ws)

    header_idx = find_type_header(rows)
    if header_idx < 0:
        logger.warning(
            "Sheet '%s': no header row with date objects — ANBIMA may not "
            "publish this sheet in the current edition; emitting nothing",
            sheet_fragment,
        )
        return []
    header = rows[header_idx]
    months, ytd_cols, twelve_m_cols = map_type_columns(header)
    if not months:
        logger.warning("Sheet '%s': header row %d has no month columns",
                       sheet_fragment, header_idx)
        return []
    anchor = max(months.values())   # latest published month: where YTD/12m land

    records: List[Dict] = []
    current_category: Optional[str] = None
    skipped = 0

    for row in rows[header_idx + 1:]:
        values: List[Tuple[str, date, float]] = []
        for col_idx, ref in months.items():
            val = _safe_float(row[col_idx]) if col_idx < len(row) else None
            if val is not None:
                values.append((monthly_metric, ref, val))
        if ytd_metric:
            for col_idx in ytd_cols:
                val = _safe_float(row[col_idx]) if col_idx < len(row) else None
                if val is not None:
                    values.append((ytd_metric, anchor, val))
        if twelvem_metric:
            for col_idx in twelve_m_cols:
                val = _safe_float(row[col_idx]) if col_idx < len(row) else None
                if val is not None:
                    values.append((twelvem_metric, anchor, val))

        classified = classify_type_row(row, current_category, bool(values))
        if classified is None:
            skipped += 1
            continue
        level, type_id, type_name, category = classified

        if level == LEVEL_CATEGORY and category != TOTAL_CATEGORY:
            current_category = category

        for metric, ref, val in values:
            records.append(_record(
                ref, category, type_id, type_name, metric, val,
                level, source_sheet_label, boletim_ref,
            ))

    logger.debug("Sheet '%s': %d records, %d rows skipped",
                 sheet_fragment, len(records), skipped)
    return records


def parse_pl_tipo(wb: openpyxl.Workbook, boletim_ref: str) -> List[Dict]:
    """Pág. 5 - PL por Tipo → pl_brl_mm for every class and type."""
    return parse_type_sheet(
        wb, "Pág. 5", "Pág. 5 - PL por Tipo",
        monthly_metric=METRIC_PL,
        ytd_metric="",          # the PL sheet publishes no YTD / 12m columns
        twelvem_metric="",
        boletim_ref=boletim_ref,
    )


def parse_capliq_tipo(wb: openpyxl.Workbook, boletim_ref: str) -> List[Dict]:
    """Pág. 9 - Cap. Líq. por Tipo → captacao_liquida_* for every class and type."""
    return parse_type_sheet(
        wb, "Pág. 9", "Pág. 9 - Cap. Líq. por Tipo",
        monthly_metric=METRIC_CAPLIQ,
        ytd_metric=METRIC_CAPLIQ_YTD,
        twelvem_metric=METRIC_CAPLIQ_12M,
        boletim_ref=boletim_ref,
    )


def parse_rentabilidade(wb: openpyxl.Workbook, boletim_ref: str) -> List[Dict]:
    """Pág.11 - Rentabilidade por Tipo → rentabilidade_* for every type.

    This sheet is types-only in practice: ANBIMA prints the class rows as bare
    labels with no numbers, and omits ETF / FIDC / FIP / FIAGRO / FII / Off Shore
    entirely in some editions. Both are handled — an empty class row still sets
    the class its types belong to, and a missing section simply yields no rows.
    """
    return parse_type_sheet(
        wb, "Pág.11", "Pág.11 - Rentabilidade por Tipo",
        monthly_metric=METRIC_RENT,
        ytd_metric=METRIC_RENT_YTD,
        twelvem_metric=METRIC_RENT_12M,
        boletim_ref=boletim_ref,
    )


# ── assembly ──────────────────────────────────────────────────────────────────

def dedupe_records(records: List[Dict]) -> List[Dict]:
    """Collapse records that share a primary key, keeping the FIRST.

    The class sheets and the type sheets both publish the class aggregates, so
    e.g. ETF's July PL arrives from Pág. 4 and again from Pág. 5. The class
    sheets carry the full history and are parsed first, so they win. A genuine
    disagreement between two sheets is a data-quality signal — it is logged, not
    silently averaged or overwritten.
    """
    seen: Dict[Tuple, Dict] = {}
    out: List[Dict] = []
    for rec in records:
        key = (rec["reference_date"], rec["anbima_category"],
               rec["anbima_type_name"], rec["metric"], rec["level"])
        kept = seen.get(key)
        if kept is None:
            seen[key] = rec
            out.append(rec)
            continue
        a, b = kept["value"], rec["value"]
        if abs(a - b) > max(1e-6, abs(a) * 1e-6):
            logger.warning(
                "ANBIMA sheets disagree for %s: %s=%s (%s) vs %s (%s) — keeping the first",
                key, kept["source_sheet"], a, kept["source_sheet"], b, rec["source_sheet"],
            )
    if len(out) < len(records):
        logger.info("ANBIMA: %d duplicate keys collapsed (%d → %d records)",
                    len(records) - len(out), len(records), len(out))
    return out


def parse_boletim(xlsx_bytes: bytes, boletim_ref: str) -> List[Dict]:
    """Parse every relevant sheet and return records ready for upsert."""
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)

    records: List[Dict] = []
    # Class sheets first: they hold the full history, so they win any dedupe
    # against the type sheets' ~19-month window.
    parsers = [
        ("Pág. 4 (PL classe)",      parse_pl_classe),
        ("Pág. 8 (CapLiq classe)",  parse_capliq_classe),
        ("Pág. 13 (N° fundos)",     parse_nfundos),
        ("Pág. 5 (PL tipo)",        parse_pl_tipo),
        ("Pág. 9 (CapLiq tipo)",    parse_capliq_tipo),
        ("Pág. 11 (Rentabilidade)", parse_rentabilidade),
    ]
    for label, fn in parsers:
        try:
            batch = fn(wb, boletim_ref)
            logger.info("  %s → %d records", label, len(batch))
            records.extend(batch)
        except Exception as exc:
            logger.warning("Parser failed for %s: %s", label, exc)

    wb.close()
    return dedupe_records(records)


# ── ANBIMA API ────────────────────────────────────────────────────────────────

async def fetch_latest_boletim_url() -> Tuple[str, str]:
    """
    Query the ANBIMA Strapi API and return (download_url, boletim_ref).
    boletim_ref = the upload filename slug (used for provenance).
    """
    url = STRAPI_BASE + STRAPI_API_PATH
    async with httpx.AsyncClient(timeout=30) as client:
        resp = await client.get(url)
        resp.raise_for_status()
    data = resp.json()
    try:
        rel_path = (
            data["data"][0]["attributes"]["template"]["attachment"]["data"]["attributes"]["url"]
        )
    except (KeyError, IndexError, TypeError) as exc:
        raise RuntimeError(f"Unexpected ANBIMA Strapi response structure: {exc}") from exc
    download_url = STRAPI_BASE + rel_path
    boletim_ref  = rel_path.split("/")[-1]   # e.g. "Anexo_boletim_fundos_investimento_abril_valor_7e6dc82403.xlsx"
    return download_url, boletim_ref


async def download_xlsx(url: str) -> bytes:
    """Download the XLSX file and return raw bytes."""
    async with httpx.AsyncClient(timeout=120, follow_redirects=True) as client:
        resp = await client.get(url)
        resp.raise_for_status()
    return resp.content


# ── ingestor class ────────────────────────────────────────────────────────────

class AnbimaIngestor:
    """
    Fetches and upserts the ANBIMA boletim's class + type metrics.

    Public interface (mirrors other pipeline classes in this repo):
      daily_update()  — fetch latest boletim and upsert; idempotent
      backfill()      — not applicable here (historical data is embedded in
                        every boletim XLSX); daily_update() already covers
                        full history from each file.
    """

    def __init__(self):
        self._pg = get_pg_client()

    # ── internal helpers ──────────────────────────────────────────────────────

    # Audit rows go through src/pipeline/ingest_log — the one writer, so the
    # 2026-07-25 incident (these helpers once sent `notes`/`error_message`,
    # columns that do not exist; _log_start raised before the upsert try/except
    # and ANBIMA silently skipped every day) cannot recur here or elsewhere.
    # Provenance for the source file is not lost: every record carries boletim_ref.

    def _log_start(self, conn, run_id: str, boletim_ref: str) -> None:
        logger.info("[anbima] run %s boletim=%s", run_id, boletim_ref)
        ingest_log.start(conn, run_id, LOG_ENTITY, LOG_DOC_TYPE, upsert=upsert_rows)

    def _log_finish(
        self,
        conn,
        run_id: str,
        status: str,
        rows_upserted: int,
        error: Optional[str] = None,
    ) -> None:
        ingest_log.finish(conn, run_id, LOG_ENTITY, LOG_DOC_TYPE,
                          status=status, rows=rows_upserted, error=error,
                          upsert=upsert_rows)

    async def daily_update(self) -> Dict[str, int]:
        """
        Fetch the latest published ANBIMA boletim and upsert every class metric.
        Idempotent: re-running with the same boletim inserts/updates the same rows.

        Returns: {'anbima_etf': <rows_upserted>}  (key kept for the daily-run
        totals dict and the log entity, both of which predate the widening)
        """
        run_id = str(uuid.uuid4())
        rows_upserted = 0

        logger.info("[anbima] Starting daily update (run_id=%s)", run_id)
        download_url, boletim_ref = await fetch_latest_boletim_url()
        logger.info("[anbima] Latest boletim: %s", boletim_ref)

        xlsx_bytes = await download_xlsx(download_url)
        logger.info("[anbima] Downloaded %d bytes", len(xlsx_bytes))

        records = parse_boletim(xlsx_bytes, boletim_ref)
        logger.info("[anbima] Parsed %d records total", len(records))

        if not records:
            logger.warning("[anbima] No records parsed — skipping upsert")
            return {LOG_ENTITY: 0}

        self._log_start(self._pg, run_id, boletim_ref)
        try:
            upsert_rows(
                self._pg,
                TABLE,
                records,
                conflict_columns=CONFLICT_COLUMNS,
            )
            rows_upserted = len(records)
            # 'ok' (not 'success') — staleness/coverage checks count status='ok'.
            self._log_finish(self._pg, run_id, "ok", rows_upserted)
            logger.info("[anbima] Upserted %d rows", rows_upserted)
        except Exception as exc:
            self._log_finish(self._pg, run_id, "error", 0, ingest_log.describe(exc))
            logger.error("[anbima] Upsert failed: %s", exc)
            raise

        return {LOG_ENTITY: rows_upserted}

    async def backfill(self, start_year: int = 2006) -> Dict[str, int]:
        """Backfill ANBIMA class metrics.

        This delegates to daily_update() *by design*, not as a stub:

        - The ANBIMA Strapi endpoint exposes ONLY the latest boletim (sorted by
          publishedAt desc, pageSize=1) — there is no list/archive/date-filter
          parameter to enumerate or fetch older files.
        - However, each boletim XLSX already embeds the FULL monthly history back
          to 2006 in the class-level sheets (Pág.4 PL, Pág.8 captação, Pág.13 nº
          fundos). So a single fetch backfills all class-level series — there is
          nothing more to retrieve.

        Caveat: the type-level sheets (Pág.5/9/11) carry only a rolling ~19-month
        window. Deep type-level history is genuinely unavailable from this API;
        obtaining it would require a separate source of archived boletim files
        (out of scope here).
        """
        logger.info(
            "[anbima] backfill() called — delegating to daily_update() "
            "(boletim already contains full history back to %d)", start_year
        )
        return await self.daily_update()


# Backwards-compatible alias for the ETF-only era of this module.
AnbimaEtfIngestor = AnbimaIngestor


# ── CLI entry point ───────────────────────────────────────────────────────────

if __name__ == "__main__":
    import argparse

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)-8s %(name)s: %(message)s",
    )

    parser = argparse.ArgumentParser(description="ANBIMA boletim ingestor")
    parser.add_argument(
        "--mode",
        choices=["daily", "backfill"],
        default="daily",
        help="daily: fetch latest boletim; backfill: alias for daily (full history embedded)",
    )
    args = parser.parse_args()

    ingestor = AnbimaIngestor()
    if args.mode == "backfill":
        result = asyncio.run(ingestor.backfill())
    else:
        result = asyncio.run(ingestor.daily_update())
    print(result)
